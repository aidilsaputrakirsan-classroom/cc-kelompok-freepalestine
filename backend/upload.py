"""
Upload module — Parser CSV/XLSX untuk import data ke SalesData atau InboxItem.

Skema kolom STRICT:
- Revenue/Sales (Menu Revenue Analytics) wajib mengandung kolom berikut:
    witel, channel, product,
    revenue_target, revenue_actual,
    sales_target, sales_actual,
    period_month, period_year
  (alias yang diterima dijelaskan di bawah.)

- Inbox (Menu Customer Care & NPS) wajib mengandung kolom berikut:
    title, witel, status, priority
  (kolom opsional: description, category, assigned_to)

File yang tidak memenuhi skema ini akan ditolak (ValueError) — tidak akan
pernah tersimpan ke database.
"""
import io
import csv
from typing import List, Tuple, Optional
from openpyxl import load_workbook


# ============================================================
# KONSTANTA VALIDASI
# ============================================================

ALLOWED_WITELS = {"BALIKPAPAN", "KALBAR", "KALSELTENG", "KALTIMTARA"}
ALLOWED_STATUSES = {"pending", "in_progress", "completed", "rejected"}
ALLOWED_PRIORITIES = {"low", "medium", "high", "critical"}
ALLOWED_PRODUCTS = {"HSI", "B2B", "WMS"}

# Alias kolom — frontend form menggunakan nama kiri,
# tapi file CSV/XLSX boleh pakai salah satu nama di kanan.
SALES_COLUMNS = {
    "witel":           ["witel", "witel_billing"],
    "channel":         ["channel"],
    "product":         ["product", "produk"],
    "revenue_target":  ["revenue_target", "target_revenue"],
    "revenue_actual":  ["revenue_actual", "actual_revenue"],
    "sales_target":    ["sales_target", "ssl_target"],
    "sales_actual":    ["sales_actual", "ssl_actual"],
    "period_month":    ["period_month"],
    "period_year":     ["period_year"],
}

# Semua kolom sales bersifat wajib (agar integritas data Revenue
# Analytics terjaga dan kolom tabel terisi penuh).
SALES_REQUIRED = list(SALES_COLUMNS.keys())

SALES_OPTIONAL = {
    "telda":          ["telda", "telkom_daerah"],
    "flag":           ["flag"],
    "flag_2":         ["flag_2", "flag2"],
    "flag_3":         ["flag_3", "flag3"],
    "nama_pelanggan": ["nama_pelanggan", "pelanggan"],
    "layanan":        ["layanan", "layanan_produk"],
    "nama_am":        ["nama_am", "am", "account_manager"],
}

INBOX_COLUMNS = {
    "title":    ["title", "judul"],
    "witel":    ["witel", "witel_billing"],
    "status":   ["status"],
    "priority": ["priority", "prioritas"],
}
INBOX_REQUIRED = list(INBOX_COLUMNS.keys())

INBOX_OPTIONAL = {
    "description": ["description", "deskripsi", "keterangan"],
    "category":    ["category", "kategori"],
    "assigned_to": ["assigned_to", "assigned", "teknisi"],
}

# Label ramah pengguna — dipakai di pesan error supaya user tidak
# melihat nama kolom teknis seperti "revenue_target" atau "period_year".
COLUMN_LABELS = {
    "witel":          "Witel",
    "channel":        "Channel",
    "product":        "Produk",
    "revenue_target": "Target Revenue",
    "revenue_actual": "Realisasi Revenue",
    "sales_target":   "Target SSL",
    "sales_actual":   "Realisasi SSL",
    "period_month":   "Bulan",
    "period_year":    "Tahun",
    "title":          "Judul Tiket",
    "status":         "Status",
    "priority":       "Prioritas",
}


# ============================================================
# PARSER FILE
# ============================================================

def parse_file(file_bytes: bytes, filename: str) -> List[dict]:
    """Parse CSV atau XLSX → list of row dicts (header sebagai key)."""
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext == "csv":
        return _parse_csv(file_bytes)
    if ext in ("xlsx", "xls"):
        return _parse_xlsx(file_bytes)
    raise ValueError(
        f"Format file tidak didukung: .{ext}. Gunakan .csv atau .xlsx."
    )


def _parse_csv(file_bytes: bytes) -> List[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_xlsx(file_bytes: bytes) -> List[dict]:
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
    return out


# ============================================================
# HELPER
# ============================================================

def _norm(s) -> str:
    """Normalisasi nama kolom: lowercase, strip, spasi/dash → underscore."""
    if s is None:
        return ""
    return str(s).strip().lower().replace(" ", "_").replace("-", "_")


def _build_col_map(row_keys, column_aliases: dict) -> dict:
    """
    Untuk setiap kolom logis (key di column_aliases), cari header aktual
    di file yang cocok dengan salah satu alias.
    Return: {logical_name: actual_header_in_file or None}.
    """
    norm_to_actual = {_norm(k): k for k in row_keys if k is not None}
    col_map = {}
    for logical, aliases in column_aliases.items():
        found = None
        for alias in aliases:
            if _norm(alias) in norm_to_actual:
                found = norm_to_actual[_norm(alias)]
                break
        col_map[logical] = found
    return col_map


def _safe_float(val, default=0.0):
    if val is None or val == "":
        return default
    try:
        s = str(val).replace(",", "").replace("(", "-").replace(")", "").strip()
        if s == "":
            return default
        return float(s)
    except (ValueError, TypeError):
        raise ValueError(f"Nilai angka tidak valid: '{val}'")


def _safe_int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(_safe_float(val, default))
    except ValueError:
        raise


def _safe_str(val, default: Optional[str] = ""):
    if val is None:
        return default
    s = str(val).strip()
    return s if s else default


def _normalize_witel(val: str) -> str:
    """Ekstrak nama Witel yang valid dari input (BALIKPAPAN, KALBAR, dll)."""
    s = _safe_str(val).upper()
    for w in ALLOWED_WITELS:
        if w in s:
            return w
    return s  # akan divalidasi di caller


# ============================================================
# MAPPER — SALES / REVENUE ANALYTICS
# ============================================================

def map_sales_rows(rows: List[dict]) -> Tuple[List[dict], List[str]]:
    """
    Map rows menjadi format SalesData. Validasi ketat:
    - Semua kolom di SALES_REQUIRED harus hadir di header.
    - Baris yang memiliki nilai invalid (witel/product di luar enum,
      bulan di luar 1-12, dll) akan dikumpulkan sebagai error.
    - Jika tidak ada satu pun baris valid, caller akan menolak upload.
    """
    if not rows:
        raise ValueError("File kosong: tidak ada baris data yang bisa diproses.")

    header_keys = list(rows[0].keys())
    col_map = _build_col_map(header_keys, SALES_COLUMNS)
    opt_map = _build_col_map(header_keys, SALES_OPTIONAL)

    missing = [logical for logical in SALES_REQUIRED if col_map[logical] is None]
    if missing:
        pretty = [COLUMN_LABELS.get(c, c) for c in missing]
        raise ValueError(
            "File ditolak: format tidak sesuai skema Revenue Analytics. "
            f"Kolom wajib yang hilang: {', '.join(pretty)}."
        )

    mapped: List[dict] = []
    errors: List[str] = []

    for i, row in enumerate(rows):
        line = i + 2  # +1 untuk header, +1 untuk 1-indexed
        try:
            witel = _normalize_witel(row.get(col_map["witel"]))
            if witel not in ALLOWED_WITELS:
                raise ValueError(
                    f"witel '{row.get(col_map['witel'])}' tidak dikenal. "
                    f"Gunakan salah satu: {', '.join(sorted(ALLOWED_WITELS))}."
                )

            channel = _safe_str(row.get(col_map["channel"]))
            if not channel:
                raise ValueError("channel wajib diisi.")

            product = _safe_str(row.get(col_map["product"])).upper()
            if product not in ALLOWED_PRODUCTS:
                raise ValueError(
                    f"product '{product}' tidak dikenal. "
                    f"Gunakan salah satu: {', '.join(sorted(ALLOWED_PRODUCTS))}."
                )

            revenue_target = _safe_float(row.get(col_map["revenue_target"]))
            revenue_actual = _safe_float(row.get(col_map["revenue_actual"]))
            sales_target = _safe_int(row.get(col_map["sales_target"]))
            sales_actual = _safe_int(row.get(col_map["sales_actual"]))

            period_month = _safe_int(row.get(col_map["period_month"]))
            period_year = _safe_int(row.get(col_map["period_year"]))

            if not (1 <= period_month <= 12):
                raise ValueError(f"period_month '{period_month}' harus 1-12.")
            if not (2020 <= period_year <= 2030):
                raise ValueError(f"period_year '{period_year}' harus 2020-2030.")

            if any(v < 0 for v in (revenue_target, revenue_actual, sales_target, sales_actual)):
                raise ValueError("Nilai target/actual tidak boleh negatif.")

            mapped.append({
                "witel":          witel,
                "telda":          _safe_str(row.get(opt_map["telda"])) or None if opt_map["telda"] else None,
                "channel":        channel,
                "product":        product,
                "revenue_target": revenue_target,
                "revenue_actual": revenue_actual,
                "sales_target":   sales_target,
                "sales_actual":   sales_actual,
                "period_month":   period_month,
                "period_year":    period_year,
                "flag":           _safe_str(row.get(opt_map["flag"])) or None if opt_map["flag"] else None,
                "flag_2":         _safe_str(row.get(opt_map["flag_2"])) or None if opt_map["flag_2"] else None,
                "flag_3":         _safe_str(row.get(opt_map["flag_3"])) or None if opt_map["flag_3"] else None,
                "nama_pelanggan": _safe_str(row.get(opt_map["nama_pelanggan"])) or None if opt_map["nama_pelanggan"] else None,
                "layanan":        _safe_str(row.get(opt_map["layanan"])) or None if opt_map["layanan"] else None,
                "nama_am":        _safe_str(row.get(opt_map["nama_am"])) or None if opt_map["nama_am"] else None,
            })
        except Exception as e:
            errors.append(f"Baris {line}: {e}")

    return mapped, errors


# ============================================================
# MAPPER — INBOX / CUSTOMER CARE
# ============================================================

def map_inbox_rows(rows: List[dict]) -> Tuple[List[dict], List[str]]:
    """
    Map rows menjadi format InboxItem. Validasi ketat:
    - Kolom wajib: title, witel, status, priority (header harus ada).
    - Setiap baris divalidasi: witel/status/priority harus ∈ enum.
    """
    if not rows:
        raise ValueError("File kosong: tidak ada baris data yang bisa diproses.")

    header_keys = list(rows[0].keys())
    col_map = _build_col_map(header_keys, INBOX_COLUMNS)
    opt_map = _build_col_map(header_keys, INBOX_OPTIONAL)

    missing = [logical for logical in INBOX_REQUIRED if col_map[logical] is None]
    if missing:
        pretty = [COLUMN_LABELS.get(c, c) for c in missing]
        raise ValueError(
            "File ditolak: format tidak sesuai skema Customer Care & NPS. "
            f"Kolom wajib yang hilang: {', '.join(pretty)}."
        )

    mapped: List[dict] = []
    errors: List[str] = []

    for i, row in enumerate(rows):
        line = i + 2
        try:
            title = _safe_str(row.get(col_map["title"]))
            if not title:
                raise ValueError("title wajib diisi.")

            witel = _normalize_witel(row.get(col_map["witel"]))
            if witel not in ALLOWED_WITELS:
                raise ValueError(
                    f"witel '{row.get(col_map['witel'])}' tidak dikenal. "
                    f"Gunakan salah satu: {', '.join(sorted(ALLOWED_WITELS))}."
                )

            status = _safe_str(row.get(col_map["status"])).lower()
            if status not in ALLOWED_STATUSES:
                raise ValueError(
                    f"status '{status}' tidak dikenal. "
                    f"Gunakan salah satu: {', '.join(sorted(ALLOWED_STATUSES))}."
                )

            priority = _safe_str(row.get(col_map["priority"])).lower()
            if priority not in ALLOWED_PRIORITIES:
                raise ValueError(
                    f"priority '{priority}' tidak dikenal. "
                    f"Gunakan salah satu: {', '.join(sorted(ALLOWED_PRIORITIES))}."
                )

            mapped.append({
                "title":       title,
                "description": _safe_str(row.get(opt_map["description"])) or None if opt_map["description"] else None,
                "status":      status,
                "priority":    priority,
                "witel":       witel,
                "category":    _safe_str(row.get(opt_map["category"])) or None if opt_map["category"] else None,
                "assigned_to": _safe_str(row.get(opt_map["assigned_to"])) or None if opt_map["assigned_to"] else None,
            })
        except Exception as e:
            errors.append(f"Baris {line}: {e}")

    return mapped, errors
